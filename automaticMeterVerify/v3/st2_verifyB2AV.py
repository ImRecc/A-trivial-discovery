# -*- coding: utf-8 -*-
import json
import time
import requests
import base64
import re
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import urllib3
import sys
from difflib import SequenceMatcher

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================= 配置区 =================
ALI_APP_CODE = "10d5fd0f043b408ca829a41a7a49d6bd"
ALI_ENDPOINT = "https://gas.market.alicloudapi.com/api/predict/gas_meter_end2end"

TASK_FILE = "todo_tasks.json"
RESULT_FILE = "audit_report_final.xlsx"
CACHE_FILE = "processed_cache_aliyun.json" 
# ==========================================

def call_aliyun_meter_ocr(img_path, app_code):
    """调用阿里云 OCR"""
    try:
        with open(img_path, "rb") as f:
            img_data = base64.b64encode(f.read()).decode('utf-8')
        
        headers = {
            'Authorization': f'APPCODE {app_code}',
            'Content-Type': 'application/json; charset=UTF-8'
        }
        response = requests.post(ALI_ENDPOINT, headers=headers, json={"image": img_data}, verify=False, timeout=20)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f" API报错: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f" 网络异常: {e}")
        return None

def clean_to_int(val):
    if not val: return "0"
    s = str(val).upper().replace('U', '0').replace('O', '0').replace('Z', '2')
    if '.' in s: s = s.split('.')[0]
    res = re.sub(r"\D", "", s).lstrip('0')
    return res if res else "0"

def extract_best_number(ocr_result, target_int):
    if not ocr_result or not isinstance(ocr_result, dict): return "未识别"
    candidates = set()
    
    def recurse(obj):
        if isinstance(obj, dict):
            for v in obj.values(): recurse(v)
        elif isinstance(obj, list):
            for item in obj: recurse(item)
        elif isinstance(obj, (str, int, float)):
            c = clean_to_int(str(obj))
            if 2 <= len(c) <= 8: candidates.add(c)
    
    recurse(ocr_result)
    if not candidates: return "未识别到数字"

    for c in candidates:
        if target_int in c: return c # 只要包含就认为是这张图里的主要数字
    
    return max(candidates, key=lambda c: SequenceMatcher(None, target_int, c).ratio())

def load_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except: return {}
    return {}

def save_cache(cache_obj):
    """持久化保存缓存"""
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache_obj, f, ensure_ascii=False, indent=2)

def append_to_excel(row_data, result_file):
    """追加写入 Excel"""
    if not os.path.exists(result_file):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "核对报告"
        sheet.append(["表号", "用户名", "期望整数", "AI识别整数", "匹配情况", "图片路径"])
    else:
        wb = load_workbook(result_file)
        sheet = wb.active

    sheet.append(row_data)
    curr_row = sheet.max_row
    res = row_data[4]
    
    # 颜色填充
    fill = None
    if res == "完全不匹配":
        fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    elif res == "包含匹配":
        fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    
    if fill:
        sheet.cell(row=curr_row, column=5).fill = fill
    wb.save(result_file)

def main():
    print("=== 阿里云 OCR 审计系统 (实时同步版) ===")
    
    # 1. 加载任务
    if not os.path.exists(TASK_FILE):
        print(f"找不到任务文件 {TASK_FILE}")
        return
    with open(TASK_FILE, 'r', encoding='utf-8') as f:
        tasks = json.load(f)

    # 2. 加载缓存
    cache = load_cache()
    
    # 统计一下
    todo_tasks = [t for t in tasks if str(t.get('meter_id')) not in cache]
    print(f"总任务: {len(tasks)} 条")
    print(f"已跳过 (缓存): {len(tasks) - len(todo_tasks)} 条")
    print(f"待执行: {len(todo_tasks)} 条")
    print("-" * 30)

    processed_now = 0
    for index, task in enumerate(tasks):
        m_id = str(task.get("meter_id"))
        name = task.get("customer_name", "未知")
        
        # 核心跳过逻辑：检查内存中的 cache 字典
        if m_id in cache:
            continue

        target_int = clean_to_int(task.get("target_val"))
        img_p = task.get("img_path")
        
        print(f"[{index+1}/{len(tasks)}] 处理: {name} (ID: {m_id})")
        
        # 调用接口
        ocr_res = call_aliyun_meter_ocr(img_p, ALI_APP_CODE)
        
        # 处理结果
        ai_raw = extract_best_number(ocr_res, target_int) if ocr_res else "失败"
        a_int = clean_to_int(ai_raw)
        
        if target_int == a_int: match_res = "完全匹配"
        elif target_int in a_int: match_res = "包含匹配"
        else: match_res = "完全不匹配"

        # 写入 Excel
        row = [m_id, name, target_int, a_int, match_res, f'=HYPERLINK("{img_p}", "查看图片")']
        append_to_excel(row, RESULT_FILE)

        # 【关键改进】实时更新内存 cache 并保存到文件
        cache[m_id] = {"ai_num": a_int, "match": match_res}
        save_cache(cache) 
        
        processed_now += 1
        print(f"   └─ 结果: {a_int} ({match_res})")
        time.sleep(0.3)

    print(f"\n任务结束！本次新处理 {processed_now} 条记录。")

if __name__ == "__main__":
    main()