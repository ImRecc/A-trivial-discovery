# -*- coding: utf-8 -*-
import json
import time
import requests
import base64
import re
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================= 配置区 =================
# api
BAIDU_API_KEY = ""
BAIDU_SECRET_KEY = ""

TASK_FILE = "todo_tasks.json"
RESULT_FILE = "audit_report_final.xlsx"
CACHE_FILE = "processed_cache.json"

MAX_WORKERS = 8  # 不如直接拉满
# ==========================================

file_lock = threading.Lock()

def get_access_token():
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {"grant_type": "client_credentials", "client_id": BAIDU_API_KEY, "client_secret": BAIDU_SECRET_KEY}
    try:
        response = requests.get(url, params=params, verify=False, timeout=10)
        return response.json().get("access_token")
    except: return None

def call_baidu_ocr(img_path, token):
    """
    统一调用接口。
    注意：如果仪表接口(meter)返回为空，会自动尝试高精度接口(accurate_basic)
    """
    def _request(api_url):
        try:
            with open(img_path, "rb") as f:
                img_data = base64.b64encode(f.read()).decode('utf-8')
            payload = {"image": img_data}
            headers = {'Content-Type': 'application/x-www-form-urlencoded'}
            resp = requests.post(api_url, headers=headers, data=payload, verify=False, timeout=20)
            return resp.json()
        except: return None

    # 1. 先试仪表识别
    res = _request(f"https://aip.baidubce.com/rest/2.0/ocr/v1/meter?access_token={token}")
    
    # 2. 健壮性检查：如果仪表接口没认出来(没有meter_result)，立刻补一刀通用高精度
    if not res or ("meter_result" not in res and "words_result" not in res):
        res = _request(f"https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic?access_token={token}")
    
    return res

def clean_to_int(val):
    if not val: return "0"
    s = str(val).upper().replace('U', '0').replace('O', '0').replace('Z', '2').replace('S', '5')
    if '.' in s: s = s.split('.')[0]
    res = re.sub(r"\D", "", s).lstrip('0')
    return res if res else "0"

def extract_best_value(ocr_res, target_int):
    """
    智能解析提取：兼容 meter_result 和 words_result 两种格式
    """
    if not ocr_res: return "未识别"
    
    candidates = []
    
    # 路径 A: 来自仪表识别的专用字段
    m_list = ocr_res.get("meter_result", [])
    for item in m_list:
        v = item.get("value")
        if v: candidates.append(str(v))
    
    # 路径 B: 来自通用识别或历史版本字段
    w_list = ocr_res.get("words_result", [])
    for item in w_list:
        v = item.get("words")
        if v: candidates.append(str(v))
        
    if not candidates: return "未识别"

    # 筛选逻辑：优先找包含目标数字的，其次找最相似的
    processed_c = []
    for c in candidates:
        c_clean = clean_to_int(c)
        if c_clean != "0":
            if target_int in c_clean: return c_clean # 包含匹配直接命中
            processed_c.append(c_clean)
    
    if not processed_c: return "未识别"
    
    # 返回相似度最高的
    return max(processed_c, key=lambda x: SequenceMatcher(None, target_int, x).ratio())

def append_to_excel_safe(row_data):
    """追加写入 Excel，带颜色逻辑，线程安全"""
    with file_lock:
        if not os.path.exists(RESULT_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "审计报告"
            ws.append(["表号", "用户名", "期望整数", "AI识别整数", "匹配情况", "图片路径"])
        else:
            wb = load_workbook(RESULT_FILE)
            ws = wb.active

        ws.append(row_data)
        row_idx = ws.max_row
        status = row_data[4]

        # 颜色配置
        fills = {
            "完全匹配": PatternFill(start_color="90EE90", fill_type="solid"),   # 绿
            "完全不匹配": PatternFill(start_color="FF0000", fill_type="solid"), # 红
            "包含匹配": PatternFill(start_color="FFFF00", fill_type="solid"),   # 黄
            "模糊匹配(差1)": PatternFill(start_color="ADD8E6", fill_type="solid") # 蓝
        }

        if status in fills:
            ws.cell(row=row_idx, column=5).fill = fills[status]
        
        wb.save(RESULT_FILE)

def process_single_task(task, token, cache_ref):
    m_id = str(task.get("meter_id"))
    name = task.get("customer_name", "未知")
    target_raw = task.get("target_val")
    img_p = task.get("img_path")
    
    if m_id in cache_ref: return None

    target_int = clean_to_int(target_raw)
    ocr_res = call_baidu_ocr(img_p, token)
    ai_val = extract_best_value(ocr_res, target_int)
    
    a_int = clean_to_int(ai_val)
    
    # 逻辑判定
    match_res = "完全不匹配"
    if a_int == "未识别":
        match_res = "完全不匹配(未识别)"
    else:
        try:
            t_v, a_v = int(target_int), int(a_int)
            if t_v == a_v:
                match_res = "完全匹配"
            elif target_int in a_int or a_int in target_int:
                match_res = "包含匹配"
            elif target_int in a_int:
                match_res = "包含匹配"
        except: pass

    # 写入文件
    row = [m_id, name, target_int, a_int, match_res, f'=HYPERLINK("{img_p}", "查看图片")']
    append_to_excel_safe(row)
    
    # 更新缓存
    with file_lock:
        cache_ref[m_id] = {"ai": a_int, "res": match_res}
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_ref, f, ensure_ascii=False)
            
    return f"完成: {name} -> {a_int} [{match_res}]"

def main():
    print(f"--- 百度云多线程审计系统 v2.0 (线程: {MAX_WORKERS}) ---")
    token = get_access_token()
    if not token:
        print("Token 获取失败，请检查 Key！")
        return

    if not os.path.exists(TASK_FILE): return
    with open(TASK_FILE, 'r', encoding='utf-8') as f:
        tasks = json.load(f)

    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            cache = json.load(f)
    else: cache = {}

    todo = [t for t in tasks if str(t.get('meter_id')) not in cache]
    print(f"待处理: {len(todo)} | 已跳过: {len(tasks)-len(todo)}")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_single_task, t, token, cache): t for t in todo}
        for future in as_completed(futures):
            res = future.result()
            if res: print(res)

    print(f"\n任务全部完成！结果已追加至: {RESULT_FILE}")

if __name__ == "__main__":
    main()
