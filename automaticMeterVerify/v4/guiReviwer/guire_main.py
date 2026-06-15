import sys
import os
import re
import openpyxl
from openpyxl.styles import PatternFill

from PySide6.QtCore import QFile, Qt, QEvent
from PySide6.QtGui import QPixmap
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog
from ui_audit import Ui_MainWindow

UI_FILE = "water_meter r2.ui"


class AuditApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # 1. Load UI
        self.ui=Ui_MainWindow()
        self.ui.setupUi(self)

        self.setWindowTitle("GuiRe")
        self.resize(960, 1080)  # Set default resolution to 1280x720 (16:9)

        # 2. Initialize Data
        self.tasks = []
        self.current_page = 0

        # Excel Colors (openpyxl)
        self.fill_red = PatternFill(start_color="FF4444", fill_type="solid")
        self.fill_blue = PatternFill(start_color="00B0F0", fill_type="solid")

        # 3. Install Event Filters (窃听器)
        for i in range(4):
            lbl_img = getattr(self.ui, f"img_{i}")
            lbl_img.installEventFilter(self)

        # Inside __init__, replace self.scan_files() with:
        self.open_file_dialog()
        # 4. Scan Excel files and Start
        #self.scan_files()
        self.refresh_ui()

    def scan_file(self, file_path):
        """扫描当前目录下的所有Excel"""
        self.ui.progress.setText(f"Scanning Excel file...{file_path}")
        QApplication.processEvents()  # 强制刷新UI防止卡死

        #files = [f for f in os.listdir('.') if f.lower().endswith('.xlsx') and not f.startswith('~$')]

        filename = file_path.split("/")[-1]
        #for f in files:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)  # data_only=False to read HYPERLINK formula
            ws = wb.active
            for r in range(2, ws.max_row + 1):
                # Column 5 is Status
                st = str(ws.cell(row=r, column=5).value or "")
                if "不匹配" in st or "失败" in st:
                    #this only review "不匹配"\"失败"
                    #也就是说被手改过的就不会再出现
                    # Column 6 is Image Link (图片路径)
                    cv = ws.cell(row=r, column=6).value
                    # Regex to extract path from HYPERLINK("path", "name")
                    #=HYPERLINK("C:\img.jpg", "View").
                    # This regex extracts exactly what is inside the first quotes.
                    m = re.search(r'HYPERLINK\("([^"]+)"', str(cv or ""), re.IGNORECASE)
                    img_path = m.group(1) if m else str(cv or "")

                    user_info = str(ws.cell(row=r, column=1).value or "未知")
                    target_val = str(ws.cell(row=r, column=3).value or "")

                    if img_path:
                        self.tasks.append({
                            "filename": filename,
                            "row_idx": r,
                             "target": target_val,
                            "img_path": img_path.strip(),
                            "user_info": user_info,
                            "state": 0  # 0: Default, 1: Red (Error), 2: Blue (Blurry)
                        })
            wb.close()
        except Exception as e:
            print(f"Error reading {f}: {e}")

        if not self.tasks:
            QMessageBox.information(self, "Done", "No unmatched tasks found!")
            sys.exit(0)

    def refresh_ui(self):
        """Update the 4 grids based on current_page"""
        start_idx = self.current_page * 4

        # Update top status (显示当前文件名)
        current_fn = self.tasks[start_idx]['filename'] if start_idx < len(self.tasks) else "None"
        self.setWindowTitle(f"GuiRe: {current_fn}")

        for i in range(4):
            lbl_img = getattr(self.ui, f"img_{i}")
            lbl_text = getattr(self.ui, f"text_{i}")
            task_idx = start_idx + i

            if task_idx < len(self.tasks):
                task = self.tasks[task_idx]
                lbl_text.setText(f"ID: {task['user_info']} | Target: {task['target']}")

                # Load Image (处理路径斜杠问题)
                img_p = os.path.abspath(task['img_path'].replace("/", "\\"))
                pix = QPixmap(img_p)
                if not pix.isNull():
                    lbl_img.setPixmap(pix)
                else:
                    lbl_img.setText(f"Image Lost:\n{img_p}")

                # Restore Borders (恢复边框状态)
                if task['state'] == 1:
                    lbl_img.setStyleSheet("border: 5px solid red;")
                elif task['state'] == 2:
                    lbl_img.setStyleSheet("border: 5px solid blue;")
                else:
                    lbl_img.setStyleSheet("border: none;")
            else:
                lbl_text.setText("---")
                lbl_img.clear()
                lbl_img.setText("No Data")
                lbl_img.setStyleSheet("border: none;")

        total_pages = (len(self.tasks) + 3) // 4
        self.ui.progress.setText(f"Page {self.current_page + 1} / {total_pages} | Space: Save & Next | Z: Prev")

    def eventFilter(self, obj, event):
        """Mouse Clicks for Red/Blue Borders (鼠标点击标记)"""
        if event.type() == QEvent.MouseButtonPress:
            for i in range(4):
                lbl_img = getattr(self.ui, f"img_{i}")
                if obj == lbl_img:
                    task_idx = self.current_page * 4 + i
                    if task_idx < len(self.tasks):
                        current_state = self.tasks[task_idx]['state']

                        # Left Click (Red)
                        if event.button() == Qt.LeftButton:
                            if current_state == 1:  # Toggle off (再次点击取消)
                                self.tasks[task_idx]['state'] = 0
                                lbl_img.setStyleSheet("border: none;")
                            else:
                                self.tasks[task_idx]['state'] = 1
                                lbl_img.setStyleSheet("border: 5px solid red;")

                        # Right Click (Blue)
                        elif event.button() == Qt.RightButton:
                            if current_state == 2:  # Toggle off
                                self.tasks[task_idx]['state'] = 0
                                lbl_img.setStyleSheet("border: none;")
                            else:
                                self.tasks[task_idx]['state'] = 2
                                lbl_img.setStyleSheet("border: 5px solid blue;")

                    return True
        return super().eventFilter(obj, event)

    def keyPressEvent(self, event):
        """Keyboard Controls"""
        # Space: Save current page to Excel and go Next
        if event.key() == Qt.Key_Space:
            self.save_current_page_to_excel()

            if (self.current_page + 1) * 4 < len(self.tasks):
                self.current_page += 1
                self.refresh_ui()
            else:
                QMessageBox.information(self, "Finished", "All tasks completed!")
                self.close()

        # Z: Go to Previous Page (回退)
        elif event.key() == Qt.Key_Z:
            if self.current_page > 0:
                self.current_page -= 1
                self.refresh_ui()

    def save_current_page_to_excel(self):
        """Save the 4 tasks of the current page to their respective Excel files"""
        start_idx = self.current_page * 4
        page_tasks = self.tasks[start_idx: start_idx + 4]

        if not page_tasks: return

        # Group tasks by filename to avoid opening the same file multiple times
        # (按文件名分组，防止同一个Excel被反复打开关闭)
        tasks_by_file = {}
        for t in page_tasks:
            tasks_by_file.setdefault(t['filename'], []).append(t)

        for filename, tasks in tasks_by_file.items():
            try:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active

                for t in tasks:
                    r = t['row_idx']
                    state = t['state']

                    if state == 1:  # Red
                        ws.cell(row=r, column=4).value = t['target']
                        ws.cell(row=r, column=5).value = "确实报错"
                        ws.cell(row=r, column=5).fill = self.fill_red
                    elif state == 2:  # Blue
                        ws.cell(row=r, column=4).value = "0"
                        ws.cell(row=r, column=5).value = "图片模糊"
                        ws.cell(row=r, column=5).fill = self.fill_blue
                    else:  # Default (0)
                        ws.cell(row=r, column=4).value = t['target']
                        ws.cell(row=r, column=5).value = "按期望修正"
                        ws.cell(row=r, column=5).fill = PatternFill(fill_type=None)  # Clear fill

                wb.save(filename)
                wb.close()
            except Exception as e:
                print(f"Failed to save {filename}: {e}")

    def open_file_dialog(self):
        """Pop up a Windows file selector (弹出原生文件选择框)"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File to Audit",  # Window title
            "",  # Start directory (empty means current folder)
            "Excel Files (*.xlsx);;All Files (*)"  # File filters
        )

        if file_path:
            print(f"Selected file: {file_path}")
            self.scan_file(file_path)  # We will create this next
            self.refresh_ui()
        else:
            QMessageBox.warning(self, "Warning", "No file selected! Exiting...")
            sys.exit(0)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AuditApp()
    window.show()
    sys.exit(app.exec())