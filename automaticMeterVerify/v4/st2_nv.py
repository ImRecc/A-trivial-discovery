# -*- coding: utf-8 -*-
# 阶段二：英伟达 Nemotron-Nano 视觉大模型审计 (V6 个位容错版)
import json
import requests
import base64
import re
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================= 配置区 =================
K_API_KEY = ""
INVOKE_URL = "https://integrate.api.nvidia.com/v1/chat/completions"

TASK_FILE = "todo_tasks.json"        # 来自 st1_dl.py[cite: 2]
RESULT_FILE = "audit_report_nv.xlsx" # 输出报表
CACHE_FILE = "nv_cache.json"         # 进度缓存
ERROR_FILE = "errors.json"           # 报错文件

MAX_WORKERS = 3  # 推荐 5 线程平衡速度与稳定性
# ==========================================

file_lock = threading.Lock()

SYSTEM_PROMPT = "/no_think"
USER_QUERY = (
    "任务：读取水表主读数滚轮区域的数字。"
    "规则：1. 只读黑色滚轮内的白色数字；2. 忽略红色数字和表盘；3. 忽略条形码。"
    "直接输出最终数字，不要任何多余文字。"
)

def encode_media_base64(media_file):
    with open(media_file, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")
        #找英伟达偷的

def call_nvidia_api(img_path):
    """请求英伟达 VLM 接口"""
    try:
        base64_data = encode_media_base64(img_path)
        ext = os.path.splitext(img_path)[1].lower()[1:]
        #智能地从最后一个点切开，变成 ('my.photo', '.jpg')。
        #拿 [1] 就是 .jpg。
        #.lower() 变小写，[1:] 从第2个开始切片，到最后一个，得到 jpg。
        #如果用s.split('.')则无法处理 c:/my.folder/img.jpg这种切成[c:/my, folder/img, jpg]
        mime = "image/jpeg" if ext in ["jpg", "jpeg"] else f"image/{ext}"
        #multipurpose internet mail extensions
        #image/jpeg 就是告诉服务器：“我发给你的是一张 JPEG 图片”。
        #application/json 就是告诉服务器：“我发给你的是 JSON 文本”。
        payload = {
            "model": "nvidia/nemotron-nano-12b-v2-vl",
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                #给ai的洗脑包
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": USER_QUERY},
                        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{base64_data}"}}
                    ]
                }
            ],
            "max_tokens": 64, 
            "temperature": 0.1,
            "stream": False
            #流式输出，开启会像网页那样一个字一个字的蹦出来
        }

        headers = {"Authorization": f"Bearer {K_API_KEY}", "Content-Type": "application/json"}
        #OAuth 2.0 / Bearer Token。之{"Authorization": f"Bearer {KEY}"}
        resp = requests.post(INVOKE_URL, headers=headers, json=payload, timeout=30)
        #request库会自己拼接url,headrs,这样那样的
        #塞进去就行
        if resp.status_code == 200:
            return resp.json()['choices'][0]['message']['content'].strip()
            #拿到字典里['choices']的值，[0]只需要第一个回答，有时候ai会发癫
            #['choices']是个列表来的，详见rdm.md
            #['message']字典['content']的内容.strip()来掐头去尾
            #返回值见rdm.md
            #.json()把返回量弄成大字典
            #这种可以无限连，链式调用属于
        return None
    except Exception as e:
        print(f"请求异常 {img_path}: {e}")
        return None

def clean_target(val):
    """净化期望值：处理 '621.0' -> 621 的逻辑"""
    if not val: return 0
    #真值测试， 0， 空，null都会报错
    s = str(val).strip()
    #.strip()用于清洗字符串，把字符串开头和结尾的换行符 \n、回车符 \r、空格全砍掉
    #.strip()：只能删掉字符串开头和结尾的空格、换行符 \n。
    #如果字符串中间有乱码，它无能为力
    if '.' in s: s = s.split('.')[0]
    # split('.') 会切555.0成一个列表 ["555", "0"]，拿 [0] 就是 "555"。
    res = re.sub(r"\D", "", s)
    #re, regular expression,正则表达式
    #替换所有\D(非数字)的东西为空格
    #r: Raw String，叫python不要转义反除号
    return int(res) if res else 0

def clean_ai_result(val):
    """从 AI 回复中提取首串数字"""
    if not val: return -1
    match = re.search(r'\d+', str(val))
    #d+表示第一串连续的数字
    #r: Raw String，叫python不要转义反除号
    #假设字符串是 "水表读数是1234吨"，正则 r"\d+" 找到了 1234
    #正则写成 r"读数是(\d+)吨，水表号(\d+)"

    return int(match.group()) if match else -1
    #group()或group(0)代表要所有正则表达式找到的数据，1,2,3代表第1，2，3组
    

def append_to_excel_safe(row_data):
    """带颜色的表格写入逻辑"""
    #excel很麻烦的，是一堆xml，并非json这种直接拉了用的
    #所以用了workbook()库来处理
    with file_lock:
        if not os.path.exists(RESULT_FILE):
            wb = Workbook()  # 1. 在内存里凭空捏造一个空白的 Excel 文件（Workbook 就是工作簿）
            ws = wb.active
            # 2. 一个 Excel 文件里可以有很多个 Sheet（工作表，比如 Sheet1, Sheet2）。
            # wb.active 就是获取当前默认激活的那个 Sheet。
            ws.title = "英伟达审计报告"
            # 3. 把这个 Sheet 的名字改掉(Excel 底部标签卡的名字）
            ws.append(["表号", "用户名", "期望值", "AI识别值", "判定结果", "图片路径"])
        else:
            wb = load_workbook(RESULT_FILE)
            ws = wb.active
            #wb是workbook，整个表
            #ws是激活的页面worksheet

        ws.append(row_data)
        row_idx = ws.max_row
        status = row_data[4]

        # 颜色配置：个位差视作“浅绿色”通过，与完全匹配区分开
        fills = {
            "完全匹配": PatternFill(start_color="90EE90", fill_type="solid"),     # 绿
            "模糊匹配(个位差)": PatternFill(start_color="C1FFC1", fill_type="solid"), # 浅绿
            "不匹配": PatternFill(start_color="FF0000", fill_type="solid"),       # 红
            "包含匹配": PatternFill(start_color="FFFF00", fill_type="solid")      # 黄
        }
        if status in fills:
            ws.cell(row=row_idx, column=5).fill = fills[status]
        
        wb.save(RESULT_FILE)

def process_single_task(task, cache_ref):
    m_id = str(task.get("meter_id"))
    name = task.get("customer_name", "未知")
    #value = dict.get(key, defaultValue)
    target_raw = task.get("target_val") #[cite: 4]
    img_p = task.get("img_path")
    
    if m_id in cache_ref: return None

    target_int = clean_target(target_raw)
    raw_ai_out = call_nvidia_api(img_p)
    ai_int = clean_ai_result(raw_ai_out)

    # --- 核心判定逻辑升级 ---
    if ai_int == -1:
        match_res = "不匹配"
        ai_display = "未识别"

        
        with file_lock:
            errors_data = {}
            #避免没得读
            if os.path.exists(ERROR_FILE):
                with open(ERROR_FILE, 'r', encoding='utf-8') as f:
                    try:
                        errors_data = json.load(f)
                    except:
                        errors_data = {}
                    
                     
            
            errors_data[m_id]={
                "customer_name" : name, 
                "target_val" : target_int, 
                "raw_ai_out" : raw_ai_out 
            }
            with open(ERROR_FILE, 'w', encoding='utf-8') as f:
                # indent=4,缩进4格
                # 管道f
                json.dump(errors_data, f, ensure_ascii=False, indent=4)
                #open('w')会包办如果没有文件就创立文件的功能
                #a append模式也会，但问题在于json文件只允许一个大括号，很麻烦

    elif ai_int == target_int:
        match_res = "完全匹配"
        ai_display = str(ai_int)
    # 逻辑：只要除以10的商相同，说明仅个位数不同 (如 920//10 = 92, 929//10 = 92)
    elif (ai_int // 10) == (target_int // 10):
        match_res = "模糊匹配(个位差)"
        ai_display = str(ai_int)
    else:
        # 兜底：包含匹配
        str_t, str_a = str(target_int), str(ai_int)
        if str_t in str_a or str_a in str_t:
            match_res = "包含匹配"
        else:
            match_res = "不匹配"
        ai_display = str(ai_int)

    row = [m_id, name, target_int, ai_display, match_res, f'=HYPERLINK("{img_p}", "查看图片")']
    append_to_excel_safe(row)
    
    with file_lock:
        cache_ref[m_id] = ai_display
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_ref, f, ensure_ascii=False)
            
    return f"【{match_res}】 {name}: 期望 {target_int} -> 识别 {ai_display}"

def main():
    print("--- Nemotron 审计 V6 (个位容错) ---")
    
    if not os.path.exists(TASK_FILE):
        print(f"请先运行 st1 下载脚本生成 {TASK_FILE}")
        return

    with open(TASK_FILE, 'r', encoding='utf-8') as f:
        tasks = json.load(f)

    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            cache = json.load(f)
    else: cache = {}

    todo = [t for t in tasks if str(t.get('meter_id')) not in cache]
    '''人话：
    for t in task:
        if str(t.get('meter_id')) not in cache:
            todo.append(t)
    '''
    print(f"待处理: {len(todo)} / 总数: {len(tasks)}")


    #来自 concurrent.futures 库
    #threading.Thread这种，有一万个任务，就会拉一万条线程，98x3d也熬不住这一出
    #这个会一次创立固定数量的线程，分配他们
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
    # 把 todo 列表里的任务全部派发给施工队。
    # executor.submit 就是提交任务。它不会马上返回结果，而是返回一个 futures 对象（相当于“取件凭证”）
    #哦对了，future是不可数，所以futures指的是‘期货’
        ticket_to_task_map = {executor.submit(process_single_task, t, cache): t for t in todo}
        #字典推导式
        #快速拉出一个字典{"executor.submit(process_single_task, t, cache)" : t}
        #以及python先看屁股的，先看见t for t in todo
        #所以不会出现t没赋值就传进去的情况
        '''扩写：
        ticket_to_task_map = {}
        for t in todo:
            ticket = executor.submit(process_single_task, t, cache)
            ticket_to_task_map[ticket] = t
        '''
        for ticket_to_task_map in as_completed(ticket_to_task_map):
            res = ticket_to_task_map.result()
            if res: print(res)

    '''python有个上下文管理context manager
    with会尝试搞
    __enter__(self), __exit__(self, exc_type, exc_val, exc_tb)
    (type, value, traceback)
    所以缩进离开时with会去尝试关闭，  
    并非智能f.close()之类的

    '''

    print(f"\n处理完成！结果见: {RESULT_FILE}")

if __name__ == "__main__":
    main()
